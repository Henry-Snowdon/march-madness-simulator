"""
March Madness Pool Simulator v3.2
===================================
Changes from v3.1:
  1. Person-level win probability (combined across multiple brackets)
  2. Excel index page with hyperlinks and summary
  3. Bracket rename: 'The ghost of Ali Khamenei' -> 'oh! what a lovely bracket'
  4. Team colors on FONT instead of cell background
  5. Gradient color coding in 256 scenarios (Option C)
  + New sheets: Person Summary, Game Impact Per Person, All 256 Scenarios (By Person)

Usage:
    python simulation_3_2.py
    python simulation_3_2.py --sims 1000
    python simulation_3_2.py --force "49:Duke" --force "55:Michigan"
"""

try:
    import pyodbc
except ImportError:
    pyodbc = None  # Not available on Streamlit Cloud
import pandas as pd
import numpy as np
import argparse
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill
from openpyxl.formatting.rule import FormulaRule
import datetime
import warnings
warnings.filterwarnings('ignore')

# ─── CONFIG ──────────────────────────────────────────────────────────────────
SERVER   = r'localhost\SQLEXPRESS'
DATABASE = 'MarchMadness2026'
SCHEMA   = 'mm'
N_SIMS   = 1000

BRACKET_TREE = {
    57: (49, 50), 58: (51, 52), 59: (53, 54), 60: (55, 56),
    61: (57, 60), 62: (58, 59), 63: (61, 62),
}
SLOT_ROUND       = {**{s: 3 for s in range(49, 57)}, **{s: 4 for s in range(57, 61)}, 61: 5, 62: 5, 63: 6}
POINTS_PER_ROUND = {1: 10, 2: 20, 3: 40, 4: 80, 5: 160, 6: 320}
S16_SLOTS        = list(range(49, 57))
LATE_SLOTS       = [57, 58, 59, 60, 61, 62, 63]
ALL_FUTURE_SLOTS = S16_SLOTS + LATE_SLOTS

# Team font colors (bg is now white, color applied to FONT)
TEAM_COLORS = {
    "Duke":           "003087",
    "St. John's":     "CC0000",
    "Michigan State": "18453B",
    "UConn":          "000E2F",
    "Arizona":        "CC0033",
    "Arkansas":       "9D2235",
    "Texas":          "BF5700",
    "Purdue":         "CEB888",
    "Iowa":           "FFCD00",
    "Nebraska":       "E41C38",
    "Illinois":       "E84A27",
    "Houston":        "C8102E",
    "Michigan":       "FFCB05",
    "Alabama":        "9E1B32",
    "Tennessee":      "FF8200",
    "Iowa State":     "C8102E",
}

def get_team_color(team_name):
    """Returns hex font color for team (no background)."""
    return TEAM_COLORS.get(team_name, "000000")

# ─── GRADIENT COLOR ───────────────────────────────────────────────────────────
# Fixed color for exactly 0% — dark but not extreme red
ZERO_PCT_COLOR = "C00000"

def pct_to_rgb(win_pct, baseline_pct):
    """
    Diverging gradient (green/white/red):
    - Exactly 0% always gets a fixed dark red (ZERO_PCT_COLOR)
    - Distance from baseline drives intensity
    - Near baseline = white/neutral
    - Near-zero non-zero values = dark red
    """
    # Exact zero gets a fixed color so all 0% cells match perfectly
    if win_pct == 0.0:
        return ZERO_PCT_COLOR

    v        = max(0.0, min(100.0, win_pct))
    distance = v - baseline_pct
    max_dist = max(baseline_pct, 100.0 - baseline_pct, 1.0)
    norm_dist = min(abs(distance) / max_dist, 1.0)
    abs_mod   = v / 100.0
    intensity = norm_dist * 0.7 + abs_mod * 0.3
    intensity = max(0.0, min(1.0, intensity))
    if distance >= 0:
        r = int(255 - 255 * intensity)
        g = int(255 - (255 - 97) * intensity)
        b = int(255 - 255 * intensity)
    else:
        r = int(255 - (255 - 192) * intensity)
        g = int(255 - 255 * intensity)
        b = int(255 - 255 * intensity)
    return f"{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}"

def gradient_fill(win_pct, baseline_pct):
    return PatternFill("solid", start_color=pct_to_rgb(win_pct, baseline_pct))

def gradient_font_color(win_pct, baseline_pct):
    if win_pct == 0.0:
        return "FFFFFF"  # white text on dark red zero cells
    v        = max(0.0, min(100.0, win_pct))
    distance = abs(v - baseline_pct)
    max_dist = max(baseline_pct, 100.0 - baseline_pct, 1.0)
    norm     = min(distance / max_dist, 1.0)
    return "FFFFFF" if norm > 0.55 else "000000"

# ─── PERSON MAP ──────────────────────────────────────────────────────────────
def build_person_map(scores_df):
    """Build mapping of creator -> list of bracket_ids."""
    person_map = {}
    for _, row in scores_df.iterrows():
        creator = row['bracket_creator']
        if creator not in person_map:
            person_map[creator] = []
        person_map[creator].append(int(row['bracket_id']))
    return person_map

def compute_person_win_probs(total_scores, bracket_ids, person_map):
    """
    For each simulation, find which PERSON wins
    (person with highest-scoring bracket beats all others).
    Returns dict of creator -> win_probability
    """
    n_brackets, n_sims = total_scores.shape
    bid_to_idx = {bid: i for i, bid in enumerate(bracket_ids)}

    # Build person score matrix: for each sim, person score = max of their brackets
    creators = list(person_map.keys())
    n_persons = len(creators)
    person_scores = np.zeros((n_persons, n_sims))

    for pi, creator in enumerate(creators):
        bids  = person_map[creator]
        idxs  = [bid_to_idx[bid] for bid in bids if bid in bid_to_idx]
        if idxs:
            person_scores[pi] = total_scores[idxs].max(axis=0)

    # Find winner per sim
    win_counts = np.zeros(n_persons)
    max_scores = person_scores.max(axis=0)
    for i in range(n_sims):
        mask = person_scores[:, i] == max_scores[i]
        win_counts[mask] += 1.0 / mask.sum()

    return {creator: win_counts[pi] / n_sims for pi, creator in enumerate(creators)}

def compute_person_win_probs_from_bracket_results(all_results, bracket_ids, person_map):
    """
    Given all_results (256 x n_brackets) of per-bracket win probs per scenario,
    compute per-person win probs using simulation approach.
    We approximate by running per-scenario person win probs and weighting.
    Note: This is approximate since we don't have raw scores per scenario.
    Returns all_person_results (256 x n_persons) array.
    """
    creators = list(person_map.keys())
    n_persons = len(creators)
    bid_to_idx = {bid: i for i, bid in enumerate(bracket_ids)}

    # For each scenario, person win prob ≈ 1 - P(all their brackets lose)
    # Better approximation: sum of bracket win probs (overcounts ties but close enough)
    # Most accurate: use raw total_scores — but we only have win_probs per scenario
    # We'll use: person_win_prob = P(at least one bracket wins)
    # = 1 - prod(1 - P(bracket_i wins)) [assuming independence across brackets in same pool]
    # This is an approximation since brackets aren't independent but it's directionally correct

    all_person_results = np.zeros((256, n_persons))
    for pi, creator in enumerate(creators):
        bids = person_map[creator]
        idxs = [bid_to_idx[bid] for bid in bids if bid in bid_to_idx]
        if not idxs:
            continue
        bracket_probs = all_results[:, idxs]  # (256, n_brackets_for_person)
        # P(person wins) ≈ 1 - prod(1 - P(bracket wins))
        all_person_results[:, pi] = 1 - np.prod(1 - bracket_probs, axis=1)

    return all_person_results, creators

# ─── DATABASE ────────────────────────────────────────────────────────────────
def get_connection():
    return pyodbc.connect(
        f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={SERVER};DATABASE={DATABASE};Trusted_Connection=yes;"
    )

def load_data(conn):
    scores = pd.read_sql(f"""
        SELECT bc.bracket_id, bc.bracket_name, bc.bracket_creator,
               COALESCE(SUM(CASE WHEN ts.actual_winner_id = bp.predicted_winner_id
                            THEN ts.potential_points ELSE 0 END), 0) AS current_score
        FROM {SCHEMA}.Bracket_Contestants bc
        JOIN {SCHEMA}.Bracket_Predictions bp ON bc.bracket_id = bp.bracket_id
        JOIN {SCHEMA}.Tournament_Slots    ts ON bp.slot_id    = ts.slot_id
        GROUP BY bc.bracket_id, bc.bracket_name, bc.bracket_creator
    """, conn)
    # Rename bracket
    scores['bracket_name'] = scores['bracket_name'].replace(
        'The ghost of Ali Khamenei', 'oh! what a lovely bracket')
    picks = pd.read_sql(f"""
        SELECT bp.bracket_id, bp.slot_id, bp.predicted_winner_id
        FROM {SCHEMA}.Bracket_Predictions bp
        JOIN {SCHEMA}.Tournament_Slots ts ON bp.slot_id = ts.slot_id
        WHERE ts.actual_winner_id IS NULL
    """, conn)
    slots = pd.read_sql(f"""
        SELECT slot_id, round, region, team_1_id, team_2_id
        FROM {SCHEMA}.Tournament_Slots
        WHERE actual_winner_id IS NULL AND team_1_id IS NOT NULL
        ORDER BY slot_id
    """, conn)
    strengths = pd.read_sql(f"""
        SELECT t.team_id, t.team_name, ts.kenpom_net_rating
        FROM {SCHEMA}.Team_Strengths ts JOIN {SCHEMA}.Teams t ON ts.team_id = t.team_id
    """, conn)
    teams = pd.read_sql(f"SELECT team_id, team_name FROM {SCHEMA}.Teams", conn)
    return scores, picks, slots, strengths, teams

# ─── SIMULATION CORE ─────────────────────────────────────────────────────────
def win_prob(a, b, sm):
    ra = sm.get(int(a), 20.0); rb = sm.get(int(b), 20.0)
    return 1 / (1 + np.exp(-0.15 * (ra - rb)))

def simulate_vectorized(n_sims, s16_forced, sm, slots_df, forced, completed=None):
    aw = {}
    # Pre-populate aw with completed game results so LATE_SLOTS can reference them
    if completed:
        for sid, winner_id in completed.items():
            aw[sid] = np.full(n_sims, winner_id, dtype=np.int64)
    # Only iterate over S16_SLOTS to avoid processing already-known slots
    slots_df = slots_df[slots_df['slot_id'].isin(S16_SLOTS)]
    for _, row in slots_df.iterrows():
        sid = int(row['slot_id']); ta = int(row['team_1_id']); tb = int(row['team_2_id'])
        if forced and sid in forced:
            aw[sid] = np.full(n_sims, forced[sid], dtype=np.int64)
        elif sid in s16_forced:
            aw[sid] = np.full(n_sims, s16_forced[sid], dtype=np.int64)
        else:
            p = win_prob(ta, tb, sm); d = np.random.random(n_sims)
            aw[sid] = np.where(d < p, np.full(n_sims, ta, dtype=np.int64), np.full(n_sims, tb, dtype=np.int64))
    for sid in LATE_SLOTS:
        if forced and sid in forced:
            aw[sid] = np.full(n_sims, forced[sid], dtype=np.int64); continue
        f1, f2 = BRACKET_TREE[sid]; taa = aw[f1]; tba = aw[f2]
        pairs = set(zip(taa.tolist(), tba.tolist()))
        pm = {(a, b): win_prob(a, b, sm) for a, b in pairs}
        probs = np.array([pm[(a, b)] for a, b in zip(taa, tba)])
        d = np.random.random(n_sims)
        aw[sid] = np.where(d < probs, taa, tba)
    return aw

def score_all_brackets(n_sims, aw, picks_df, scores_df, bracket_ids):
    n = len(bracket_ids); bi = {bid: i for i, bid in enumerate(bracket_ids)}
    base = np.array([float(scores_df[scores_df['bracket_id'] == bid]['current_score'].iloc[0])
                     for bid in bracket_ids]).reshape(n, 1)
    fut = np.zeros((n, n_sims))
    for sid in ALL_FUTURE_SLOTS:
        pts = POINTS_PER_ROUND[SLOT_ROUND.get(sid, 3)]; sw = aw.get(sid)
        if sw is None: continue
        for _, pr in picks_df[picks_df['slot_id'] == sid].iterrows():
            bid = int(pr['bracket_id'])
            if bid not in bi: continue
            fut[bi[bid]] += (sw == int(pr['predicted_winner_id'])) * pts
    return base + fut

def compute_win_probs(ts):
    n, s = ts.shape; wc = np.zeros(n); mx = ts.max(axis=0)
    for i in range(s):
        mask = ts[:, i] == mx[i]; wc[mask] += 1.0 / mask.sum()
    return wc / s

def scenario_kenpom_prob(perm, s16_info, sm, forced=None):
    p = 1.0
    for gi, game in enumerate(s16_info):
        if forced and game['slot_id'] in forced: continue
        pt1 = win_prob(game['team_1'], game['team_2'], sm)
        p *= pt1 if perm[gi] == 0 else (1 - pt1)
    return p

# ─── FILENAME ────────────────────────────────────────────────────────────────
def build_filename(forced_outcomes, forced_names):
    if not forced_outcomes:
        return "simulation_3_2_results.xlsx"
    parts = [f"{forced_names.get(forced_outcomes[sid],'').replace(' ','_').replace(chr(39),'').replace('.','')}_slot{sid}"
             for sid in sorted(forced_outcomes.keys())]
    return f"simulation_3_2_{'_'.join(parts)}.xlsx"

# ─── EXCEL ────────────────────────────────────────────────────────────────────
def build_excel(all_results, all_permutations, scenario_probs, s16_info,
                game_impacts, scores, bracket_ids, bracket_names,
                baseline_pct, best_pct, worst_pct, team_name_map,
                forced_outcomes, forced_names, person_map,
                all_person_results, person_baseline, person_best, person_worst, creators):

    forced_outcomes = forced_outcomes or {}
    forced_names    = forced_names    or {}
    filename        = build_filename(forced_outcomes, forced_names)
    wb              = Workbook()

    HDR_FILL  = PatternFill("solid", start_color="1F4E79")
    ALT_FILL  = PatternFill("solid", start_color="EBF3FB")
    BLK_FILL  = PatternFill("solid", start_color="FFFFFF")
    GRN_FILL  = PatternFill("solid", start_color="C6EFCE")
    RED_FILL  = PatternFill("solid", start_color="FFC7CE")
    GOLD_FILL = PatternFill("solid", start_color="FFD700")
    NAVY_FILL = PatternFill("solid", start_color="1F4E79")
    thin      = Side(style="thin",   color="CCCCCC")
    medium    = Side(style="medium", color="999999")
    bdr       = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    bdr_thick = Border(left=medium, right=medium, top=medium, bottom=medium)
    sort_order        = np.argsort(-baseline_pct)
    person_sort_order = np.argsort(-person_baseline)

    def hdr(ws, r, c, v, w=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
        if w: ws.column_dimensions[get_column_letter(c)].width = w

    def cel(ws, r, c, v, fill=None, bold=False, fmt=None, color="000000", size=9):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", size=size, bold=bold, color=color)
        cell.fill = fill or BLK_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
        if fmt: cell.number_format = fmt
        return cell

    def cel_thick(ws, r, c, v, fill=None, bold=False, fmt=None, color="000000"):
        """Like cel() but with thicker borders for 256 scenario sheets."""
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", size=9, bold=bold, color=color)
        cell.fill = fill or BLK_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr_thick
        if fmt: cell.number_format = fmt
        return cell

    # ── INDEX PAGE ───────────────────────────────────────────────────────────
    ws_idx = wb.active
    ws_idx.title = "Index"
    ws_idx.column_dimensions['A'].width = 5
    ws_idx.column_dimensions['B'].width = 40
    ws_idx.column_dimensions['C'].width = 25

    # Title
    ws_idx.merge_cells('B2:C2')
    title_cell = ws_idx.cell(row=2, column=2, value="🏀 March Madness Pool Simulator")
    title_cell.font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    title_cell.fill = NAVY_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_idx.row_dimensions[2].height = 35

    # Subtitle / metadata
    now = datetime.datetime.now().strftime("%B %d, %Y %H:%M")
    meta = [
        ("Generated:", now),
        ("Methodology:", "Pure KenPom | Exhaustive 256 S16 enumeration | KenPom-weighted baseline"),
        ("Forced outcomes:", ", ".join([f"{forced_names.get(tid,'?')} (Slot {sid})"
                                        for sid, tid in forced_outcomes.items()]) if forced_outcomes else "None"),
    ]
    for ri, (label, value) in enumerate(meta, start=4):
        ws_idx.cell(row=ri, column=2, value=label).font = Font(name="Arial", size=10, bold=True)
        ws_idx.cell(row=ri, column=3, value=value).font = Font(name="Arial", size=10)
        ws_idx.row_dimensions[ri].height = 18

    # Top 3 brackets
    ws_idx.cell(row=8, column=2, value="Top 3 Brackets by Win Probability:").font = Font(name="Arial", size=10, bold=True)
    for ri, idx in enumerate(sort_order[:3], start=9):
        bid  = bracket_ids[idx]
        name = bracket_names[bid]
        pct  = round(baseline_pct[idx]*100, 2)
        ws_idx.cell(row=ri, column=2, value=f"  {ri-8}. {name}").font = Font(name="Arial", size=10)
        ws_idx.cell(row=ri, column=3, value=f"{pct}%").font = Font(name="Arial", size=10, bold=True, color="375623")
        ws_idx.row_dimensions[ri].height = 16

    # Sheet navigation
    sheet_names = [
        ("Bracket Summary",                "Individual bracket win probabilities"),
        ("Person Summary",                 "Combined win probability per person"),
        ("Swing Analysis",                 "Game impact across all brackets"),
        ("Swing Analysis (By Person)",     "Game impact by person (combined brackets)"),
        ("Game Impact Per Bracket",        "Ranked game importance per bracket"),
        ("Game Impact Per Person",         "Ranked game importance per person"),
        ("Overall Game Impact",            "Which games matter most to the pool"),
        ("All 256 Scenarios",              "Every Sweet 16 outcome (by bracket)"),
        ("All 256 Scenarios (By Person)",  "Every Sweet 16 outcome (by person)"),
    ]
    ws_idx.cell(row=13, column=2, value="Navigate to:").font = Font(name="Arial", size=11, bold=True)
    ws_idx.row_dimensions[13].height = 20
    for ri, (sname, desc) in enumerate(sheet_names, start=14):
        link_cell = ws_idx.cell(row=ri, column=2, value=sname)
        link_cell.hyperlink = f"#{sname}!A1"
        link_cell.font = Font(name="Arial", size=10, bold=True, color="0563C1", underline="single")
        ws_idx.cell(row=ri, column=3, value=desc).font = Font(name="Arial", size=10, color="595959")
        ws_idx.row_dimensions[ri].height = 16

    # ── DEFINITIONS SECTION ───────────────────────────────────────────────────
    def_start_row = 14 + len(sheet_names) + 2

    # Section header
    ws_idx.merge_cells(f'B{def_start_row}:C{def_start_row}')
    def_hdr = ws_idx.cell(row=def_start_row, column=2, value="📖 Definitions & Methodology")
    def_hdr.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    def_hdr.fill = NAVY_FILL
    def_hdr.alignment = Alignment(horizontal="left", vertical="center")
    ws_idx.row_dimensions[def_start_row].height = 22

    definitions = [
        (
            "Monte Carlo Simulation",
            "This simulator runs thousands of hypothetical tournaments simultaneously. In each simulation, "
            "every remaining game is decided by a weighted coin flip based on KenPom team ratings — stronger "
            "teams win more often but upsets still happen. The number of simulations is set by the user "
            "(between 500 and 10,000) and this many simulations are run for every possible combination of "
            "Sweet 16 results — all 256 of them — giving a total of up to 2,560,000 individual tournament "
            "simulations. After running, we count how often each bracket finishes with the highest score. "
            "That percentage is their win probability."
        ),
        (
            "Win Percentages (Game Impact & All 256 Scenarios)",
            "The win percentages in these sheets answer the question: If this specific result (or combination "
            "of results) happens, how often does this bracket win the pool? For example, a 25% win probability "
            "means that in simulations where that outcome occurred, this bracket ended up with the highest score "
            "1 in 4 times. The baseline percentage shown in column headers is the overall win probability across "
            "all possible outcomes, weighted by their KenPom likelihood — given the relevant forced parameters."
        ),
        (
            "Color Coding (All 256 Scenarios)",
            "Each cell is colored relative to that bracket's baseline win probability. Green means that scenario "
            "is better than average for that bracket — the deeper the green, the more that combination of results "
            "helps them. Red means that scenario hurts their chances — the deeper the red, the worse it is for "
            "them. White/neutral means the scenario makes little difference either way. Cells showing exactly 0% "
            "are shown in a uniform dark red as they represent scenarios where that bracket has no realistic "
            "chance of winning the pool."
        ),
        (
            "Gold Columns (Forced Outcomes)",
            "Gold highlighted cells indicate a game outcome you have manually forced — meaning you've locked in "
            "that team as the winner before running the simulation. These columns are fixed across all 256 "
            "scenarios since the result is predetermined. The swing for forced games shows N/A since there is "
            "no alternative outcome to compare against."
        ),
        (
            "KenPom Rating",
            "KenPom net rating is a measure of team strength based on adjusted offensive and defensive efficiency. "
            "The simulator uses these ratings in the Bradley-Terry model: P(A beats B) = 1 / (1 + e^(-0.15 x "
            "(KenPom_A - KenPom_B))). A rating difference of ~10 points corresponds to roughly an 82% win "
            "probability for the stronger team."
        ),
        (
            "Person vs Bracket Win Probability",
            "Participants with multiple brackets benefit from having more than one shot at winning. The Person "
            "Win % represents the probability that at least one of a person's brackets wins the pool, calculated "
            "as: 1 - (1 - P(bracket 1 wins)) x (1 - P(bracket 2 wins)) x ... This will always be >= the best "
            "individual bracket probability, with the boost depending on how complementary their picks are."
        ),
    ]

    row = def_start_row + 1
    for title, body in definitions:
        # Title row
        ws_idx.merge_cells(f'B{row}:C{row}')
        title_c = ws_idx.cell(row=row, column=2, value=title)
        title_c.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
        title_c.fill = ALT_FILL
        title_c.alignment = Alignment(horizontal="left", vertical="center")
        ws_idx.row_dimensions[row].height = 18
        row += 1
        # Body row
        ws_idx.merge_cells(f'B{row}:C{row}')
        body_c = ws_idx.cell(row=row, column=2, value=body)
        body_c.font = Font(name="Arial", size=9, color="000000")
        body_c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_idx.row_dimensions[row].height = 70
        row += 1
        # Spacer
        ws_idx.row_dimensions[row].height = 6
        row += 1

    # Set index column widths for definitions
    ws_idx.column_dimensions['B'].width = 30
    ws_idx.column_dimensions['C'].width = 80

    # ── SHEET 1: Bracket Summary ─────────────────────────────────────────────
    ws1 = wb.create_sheet("Bracket Summary")
    for c, (t, w) in enumerate([("Rank",6),("Bracket",32),("Creator",18),("Curr Pts",9),
                                  ("Baseline %",12),("Best %",10),("Worst %",10),("Range %",10)], 1):
        hdr(ws1, 1, c, t, w)
    for rank, idx in enumerate(sort_order, 1):
        bid = bracket_ids[idx]; name = bracket_names[bid]
        creator = scores[scores['bracket_id'] == bid]['bracket_creator'].iloc[0]
        pts = int(scores[scores['bracket_id'] == bid]['current_score'].iloc[0])
        fill = ALT_FILL if rank % 2 == 0 else BLK_FILL
        for c, v in enumerate([rank, name, creator, pts,
                                round(baseline_pct[idx]*100,2), round(best_pct[idx]*100,2),
                                round(worst_pct[idx]*100,2), round((best_pct[idx]-worst_pct[idx])*100,2)], 1):
            cell = cel(ws1, rank+1, c, v, fill, bold=(c<=2))
            if c >= 5: cell.number_format = '0.00"%"'
    ws1.freeze_panes = "A2"

    # ── SHEET 2: Person Summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Person Summary")
    for c, (t, w) in enumerate([("Rank",6),("Person",22),("Brackets",40),("Best Curr Pts",12),
                                  ("Person Win %",13),("Best Bracket %",13),("Boost",10),
                                  ("Best %",10),("Worst %",10)], 1):
        hdr(ws2, 1, c, t, w)

    for rank, pi in enumerate(person_sort_order, 1):
        creator  = creators[pi]
        bids     = person_map[creator]
        bid_idxs = [bracket_ids.index(bid) for bid in bids if bid in bracket_ids]
        bracket_list = ", ".join([bracket_names[bid] for bid in bids if bid in bracket_names])
        best_curr = max(int(scores[scores['bracket_id'] == bid]['current_score'].iloc[0]) for bid in bids if bid in bracket_ids)
        person_pct = round(person_baseline[pi]*100, 2)
        best_bracket_pct = round(max(baseline_pct[i] for i in bid_idxs)*100, 2) if bid_idxs else 0
        boost = round(person_pct - best_bracket_pct, 2)
        fill = ALT_FILL if rank % 2 == 0 else BLK_FILL

        cel(ws2, rank+1, 1, rank,             fill, bold=True)
        cel(ws2, rank+1, 2, creator,           fill, bold=True)
        cel(ws2, rank+1, 3, bracket_list,      fill)
        cel(ws2, rank+1, 4, best_curr,         fill)
        cel(ws2, rank+1, 5, person_pct,        fill, fmt='0.00"%"')
        cel(ws2, rank+1, 6, best_bracket_pct,  fill, fmt='0.00"%"')
        boost_color = "375623" if boost > 0 else "000000"
        cel(ws2, rank+1, 7, boost,             GRN_FILL if boost > 0.1 else fill,
            bold=(boost > 0.1), fmt='0.00"%"', color=boost_color)
        cel(ws2, rank+1, 8, round(person_best[pi]*100,2),  fill, fmt='0.00"%"')
        cel(ws2, rank+1, 9, round(person_worst[pi]*100,2), fill, fmt='0.00"%"')
    ws2.freeze_panes = "A2"

    # ── SHEET 3: Swing Analysis ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Swing Analysis")
    ws3.row_dimensions[1].height = 35
    for c, (t, w) in enumerate([("Bracket",32),("Creator",18),("Curr Pts",9),
                                  ("Baseline %",11),("Best %",9),("Worst %",9),("Range %",9)], 1):
        hdr(ws3, 1, c, t, w)
    col = 8; game_col_map = {}
    for g in game_impacts:
        forced = bool(forced_outcomes and g['slot_id'] in forced_outcomes)
        label = f"Slot {g['slot_id']}\n{g['team_1_name']} vs {g['team_2_name']}" + ("\n[FORCED]" if forced else "")
        hdr(ws3, 1, col, label, 18); hdr(ws3, 1, col+1, "Swing", 10)
        game_col_map[g['slot_id']] = col; col += 2
    for row_idx, idx in enumerate(sort_order, start=2):
        bid = bracket_ids[idx]; name = bracket_names[bid]
        creator = scores[scores['bracket_id'] == bid]['bracket_creator'].iloc[0]
        pts = int(scores[scores['bracket_id'] == bid]['current_score'].iloc[0])
        fill = ALT_FILL if row_idx % 2 == 0 else BLK_FILL
        cel(ws3, row_idx, 1, name, fill, bold=True)
        cel(ws3, row_idx, 2, creator, fill)
        cel(ws3, row_idx, 3, pts, fill)
        for c, v in enumerate([round(baseline_pct[idx]*100,2), round(best_pct[idx]*100,2),
                                round(worst_pct[idx]*100,2), round((best_pct[idx]-worst_pct[idx])*100,2)], 4):
            cel(ws3, row_idx, c, v, fill, fmt='0.00"%"')
        for g in game_impacts:
            c = game_col_map[g['slot_id']]
            forced = bool(forced_outcomes and g['slot_id'] in forced_outcomes)
            sw = None if forced else round(g['swing'][idx]*100, 2)
            winner_lbl = (forced_names.get(forced_outcomes[g['slot_id']], '') if forced
                          else (g['team_1_name'] if g['swing'][idx] > 0 else g['team_2_name']))
            t_color = get_team_color(winner_lbl)
            cel(ws3, row_idx, c,   winner_lbl, fill, bold=True, color=t_color)
            if forced:
                cel(ws3, row_idx, c+1, "N/A", fill, color="888888")
            else:
                sw_fill = GRN_FILL if sw > 0.5 else (RED_FILL if sw < -0.5 else fill)
                cel(ws3, row_idx, c+1, sw, sw_fill, bold=(abs(sw)>1), fmt='0.00"%"',
                    color="375623" if sw > 0.5 else ("9C0006" if sw < -0.5 else "000000"))
    ws3.freeze_panes = "A2"

    # ── SHEET 3b: Swing Analysis (By Person) ────────────────────────────────
    ws3b = wb.create_sheet("Swing Analysis (By Person)")
    ws3b.row_dimensions[1].height = 35
    for c, (t, w) in enumerate([("Person",22),("Brackets",35),("Best Curr Pts",12),
                                  ("Baseline %",11),("Best %",9),("Worst %",9),("Range %",9)], 1):
        hdr(ws3b, 1, c, t, w)
    col = 8; person_game_col_map = {}
    for g in game_impacts:
        forced = bool(forced_outcomes and g['slot_id'] in forced_outcomes)
        label = f"Slot {g['slot_id']}\n{g['team_1_name']} vs {g['team_2_name']}" + ("\n[FORCED]" if forced else "")
        hdr(ws3b, 1, col, label, 18); hdr(ws3b, 1, col+1, "Swing", 10)
        person_game_col_map[g['slot_id']] = col; col += 2

    bid_to_idx_map = {bid: i for i, bid in enumerate(bracket_ids)}

    for row_idx, pi in enumerate(person_sort_order, start=2):
        creator  = creators[pi]
        bids     = person_map[creator]
        bid_idxs = [bid_to_idx_map[bid] for bid in bids if bid in bid_to_idx_map]
        bracket_list = ", ".join([bracket_names[bid] for bid in bids if bid in bracket_names])
        best_curr = max(int(scores[scores['bracket_id'] == bid]['current_score'].iloc[0])
                        for bid in bids if bid in bracket_ids)
        fill = ALT_FILL if row_idx % 2 == 0 else BLK_FILL

        cel(ws3b, row_idx, 1, creator,      fill, bold=True)
        cel(ws3b, row_idx, 2, bracket_list, fill)
        cel(ws3b, row_idx, 3, best_curr,    fill)
        for c, v in enumerate([round(person_baseline[pi]*100,2),
                                round(person_best[pi]*100,2),
                                round(person_worst[pi]*100,2),
                                round((person_best[pi]-person_worst[pi])*100,2)], 4):
            cel(ws3b, row_idx, c, v, fill, fmt='0.00"%"')

        for g in game_impacts:
            c = person_game_col_map[g['slot_id']]
            forced = bool(forced_outcomes and g['slot_id'] in forced_outcomes)

            # Person combined swing: 1 - prod(1-p) for each direction
            if forced:
                sw = None
                winner_lbl = forced_names.get(forced_outcomes[g['slot_id']], '')
            else:
                p1_combined = (1 - np.prod([1 - g['avg_if_t1_wins'][i] for i in bid_idxs])) if bid_idxs else 0
                p2_combined = (1 - np.prod([1 - g['avg_if_t2_wins'][i] for i in bid_idxs])) if bid_idxs else 0
                sw = round((p1_combined - p2_combined)*100, 2)
                winner_lbl = g['team_1_name'] if sw > 0 else g['team_2_name']

            t_color = get_team_color(winner_lbl)
            cel(ws3b, row_idx, c, winner_lbl, fill, bold=True, color=t_color)
            if forced:
                cel(ws3b, row_idx, c+1, "N/A", fill, color="888888")
            else:
                sw_fill = GRN_FILL if sw > 0.5 else (RED_FILL if sw < -0.5 else fill)
                cel(ws3b, row_idx, c+1, sw, sw_fill, bold=(abs(sw)>1), fmt='0.00"%"',
                    color="375623" if sw > 0.5 else ("9C0006" if sw < -0.5 else "000000"))
    ws3b.freeze_panes = "A2"

    # ── SHEET 4: Game Impact Per Bracket ─────────────────────────────────────
    ws4 = wb.create_sheet("Game Impact Per Bracket")
    for c, (t, w) in enumerate([("Bracket",32),("Creator",18),("Slot",7),("Team 1",18),
                                  ("Team 2",18),("Win% if T1",14),("Win% if T2",14),
                                  ("Swing",10),("Want",15),("Forced?",8)], 1):
        hdr(ws4, 1, c, t, w)
    row_idx = 2
    for idx in sort_order:
        bid = bracket_ids[idx]; name = bracket_names[bid]
        creator = scores[scores['bracket_id'] == bid]['bracket_creator'].iloc[0]
        swings = [(g['slot_id'], g['team_1_name'], g['team_2_name'],
                   g['avg_if_t1_wins'][idx]*100, g['avg_if_t2_wins'][idx]*100,
                   g['swing'][idx]*100, bool(forced_outcomes and g['slot_id'] in forced_outcomes))
                  for g in game_impacts]
        swings.sort(key=lambda x: abs(x[5]), reverse=True)
        for gi, (sid, t1, t2, p1, p2, sw, forced) in enumerate(swings):
            fill = ALT_FILL if gi % 2 == 0 else BLK_FILL
            want = t1 if sw > 0 else t2
            if forced:
                sw = None; want = forced_names.get(forced_outcomes.get(sid, 0), '')
            sw_fill = GRN_FILL if (sw and abs(sw) > 2) else fill
            w_color = get_team_color(want)
            cel(ws4, row_idx, 1, name if gi == 0 else "",    fill, bold=(gi==0))
            cel(ws4, row_idx, 2, creator if gi == 0 else "",  fill)
            cel(ws4, row_idx, 3, sid,                         fill)
            cel(ws4, row_idx, 4, t1,                          fill)
            cel(ws4, row_idx, 5, t2,                          fill)
            cel(ws4, row_idx, 6, round(p1,2),                 fill, fmt='0.00"%"')
            cel(ws4, row_idx, 7, round(p2,2),                 fill, fmt='0.00"%"')
            if forced:
                cel(ws4, row_idx, 8, "N/A", fill, color="888888")
            else:
                cel(ws4, row_idx, 8, round(abs(sw),2), sw_fill, bold=True, fmt='0.00"%"')
            cel(ws4, row_idx, 9, want,                        fill, bold=True, color=w_color)
            cel(ws4, row_idx, 10, "YES" if forced else "",    GOLD_FILL if forced else fill)
            row_idx += 1
        row_idx += 1
    ws4.freeze_panes = "A2"

    # ── SHEET 5: Game Impact Per Person ───────────────────────────────────────
    ws5 = wb.create_sheet("Game Impact Per Person")
    for c, (t, w) in enumerate([("Person",22),("Brackets",35),("Slot",7),("Team 1",18),
                                  ("Team 2",18),("Win% if T1",14),("Win% if T2",14),
                                  ("Swing",10),("Want",15),("Forced?",8)], 1):
        hdr(ws5, 1, c, t, w)
    row_idx = 2
    for pi in person_sort_order:
        creator = creators[pi]
        bids    = person_map[creator]
        bracket_list = ", ".join([bracket_names[bid] for bid in bids if bid in bracket_names])

        # Compute person-level game impact: weighted avg across their brackets
        bid_idxs = [bracket_ids.index(bid) for bid in bids if bid in bracket_ids]

        person_swings = []
        for g in game_impacts:
            forced = bool(forced_outcomes and g['slot_id'] in forced_outcomes)
            # Person win% if T1 wins = avg of their bracket win%s if T1 wins
            p1 = float(np.mean([g['avg_if_t1_wins'][i] for i in bid_idxs]))*100 if bid_idxs else 0
            p2 = float(np.mean([g['avg_if_t2_wins'][i] for i in bid_idxs]))*100 if bid_idxs else 0
            # For person: use 1 - prod(1-p) to get combined probability
            p1_combined = (1 - np.prod([1 - g['avg_if_t1_wins'][i] for i in bid_idxs]))*100 if bid_idxs else 0
            p2_combined = (1 - np.prod([1 - g['avg_if_t2_wins'][i] for i in bid_idxs]))*100 if bid_idxs else 0
            sw = 0.0 if forced else round(p1_combined - p2_combined, 2)
            person_swings.append((g['slot_id'], g['team_1_name'], g['team_2_name'],
                                   round(p1_combined,2), round(p2_combined,2), sw, forced))

        person_swings.sort(key=lambda x: abs(x[5]), reverse=True)
        for gi, (sid, t1, t2, p1, p2, sw, forced) in enumerate(person_swings):
            fill = ALT_FILL if gi % 2 == 0 else BLK_FILL
            want = t1 if sw > 0 else t2
            if forced: want = forced_names.get(forced_outcomes.get(sid, 0), '')
            sw_fill = GRN_FILL if (sw and abs(sw) > 2) else fill
            w_color = get_team_color(want)
            cel(ws5, row_idx, 1, creator if gi == 0 else "",      fill, bold=(gi==0))
            cel(ws5, row_idx, 2, bracket_list if gi == 0 else "",  fill)
            cel(ws5, row_idx, 3, sid,                              fill)
            cel(ws5, row_idx, 4, t1,                               fill)
            cel(ws5, row_idx, 5, t2,                               fill)
            cel(ws5, row_idx, 6, round(p1,2),                      fill, fmt='0.00"%"')
            cel(ws5, row_idx, 7, round(p2,2),                      fill, fmt='0.00"%"')
            if forced:
                cel(ws5, row_idx, 8, "N/A", fill, color="888888")
            else:
                cel(ws5, row_idx, 8, round(abs(sw),2), sw_fill, bold=True, fmt='0.00"%"')
            cel(ws5, row_idx, 9, want,                             fill, bold=True, color=w_color)
            cel(ws5, row_idx, 10, "YES" if forced else "",         GOLD_FILL if forced else fill)
            row_idx += 1
        row_idx += 1
    ws5.freeze_panes = "A2"

    # ── SHEET 6: Overall Game Impact ─────────────────────────────────────────
    ws6 = wb.create_sheet("Overall Game Impact")
    for c, (t, w) in enumerate([("Slot",7),("Team 1",18),("Team 2",18),("Avg Swing",24),
                                  ("Max Swing",22),("# Want T1",12),("# Want T2",12),("Forced?",8)], 1):
        hdr(ws6, 1, c, t, w)
    overall = []
    for g in game_impacts:
        forced = bool(forced_outcomes and g['slot_id'] in forced_outcomes)
        swings = g['swing']*100
        overall.append((g['slot_id'], g['team_1_name'], g['team_2_name'],
                        0.0 if forced else round(float(np.abs(swings).mean()),2),
                        0.0 if forced else round(float(np.abs(swings).max()),2),
                        int((swings>0).sum()), int((swings<0).sum()), forced))
    overall.sort(key=lambda x: x[3], reverse=True)
    for row_idx, (sid, t1, t2, avg_sw, max_sw, wt1, wt2, forced) in enumerate(overall, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else BLK_FILL
        cel(ws6, row_idx, 1, sid,    fill)
        cel(ws6, row_idx, 2, t1,     fill)
        cel(ws6, row_idx, 3, t2,     fill)
        cel(ws6, row_idx, 4, avg_sw, fill, bold=True, fmt='0.00"%"')
        cel(ws6, row_idx, 5, max_sw, fill, fmt='0.00"%"')
        cel(ws6, row_idx, 6, wt1,    GRN_FILL if wt1>wt2 else fill, bold=(wt1>wt2))
        cel(ws6, row_idx, 7, wt2,    GRN_FILL if wt2>wt1 else fill, bold=(wt2>wt1))
        cel(ws6, row_idx, 8, "YES" if forced else "", GOLD_FILL if forced else fill)
    ws6.freeze_panes = "A2"

    # ── SHEET 7: All 256 Scenarios (By Bracket) ───────────────────────────────
    ws7 = wb.create_sheet("All 256 Scenarios")
    ws7.row_dimensions[1].height = 35
    hdr(ws7, 1, 1, "Scenario", 9); hdr(ws7, 1, 2, "KenPom Prob %", 13)
    for gi, g in enumerate(s16_info):
        t1 = team_name_map.get(g['team_1'],''); t2 = team_name_map.get(g['team_2'],'')
        hdr(ws7, 1, gi+3, f"Slot {g['slot_id']}\n{t1} vs {t2}", 18)
    bcs = len(s16_info)+3
    for bi, idx in enumerate(sort_order):
        hdr(ws7, 1, bcs+bi, f"{bracket_names[bracket_ids[idx]]}\n({round(baseline_pct[idx]*100,1)}% base)", 18)
    prob_order = np.argsort(-scenario_probs)
    for di, pi in enumerate(prob_order):
        perm = all_permutations[pi]; row_idx = di+2
        fill = ALT_FILL if row_idx % 2 == 0 else BLK_FILL
        cel_thick(ws7, row_idx, 1, pi+1, fill, bold=True)
        cel_thick(ws7, row_idx, 2, round(float(scenario_probs[pi])*100,4), fill, fmt='0.0000"%"')
        for gi, game in enumerate(s16_info):
            sid = game['slot_id']; forced = bool(forced_outcomes and sid in forced_outcomes)
            if forced:
                wname = forced_names.get(forced_outcomes[sid], team_name_map.get(forced_outcomes[sid],''))
                cell_fill = PatternFill("solid", start_color="FFD700")
                # Michigan has yellow font - use black on gold
                t_color = "000000" if get_team_color(wname) in ("FFCB05","FFCD00","CEB888") else get_team_color(wname)
            else:
                wid = game['team_1'] if perm[gi] == 0 else game['team_2']
                wname = team_name_map.get(wid,'')
                cell_fill = BLK_FILL
                t_color = get_team_color(wname)
            c = ws7.cell(row=row_idx, column=gi+3, value=wname)
            c.font = Font(name="Arial", size=9, bold=True, color=t_color)
            c.fill = cell_fill
            c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bdr_thick
        for bi, idx in enumerate(sort_order):
            wp = round(all_results[pi][idx]*100, 2)
            base = baseline_pct[idx]*100
            gf = gradient_fill(wp, base)
            fc = gradient_font_color(wp, base)
            cel_thick(ws7, row_idx, bcs+bi, wp, gf, fmt='0.00"%"', color=fc)
    ws7.freeze_panes = "B2"
    ws7.auto_filter.ref = f"A1:{get_column_letter(bcs+len(bracket_ids)-1)}1"
    # Row highlight on click (conditional formatting: highlight selected row)
    last_col7 = get_column_letter(bcs+len(bracket_ids)-1)
    hl_fill7  = PatternFill("solid", start_color="FFF2CC")
    ws7.conditional_formatting.add(
        f"A2:{last_col7}{257}",
        FormulaRule(formula=['ROW()=CELL("row")'], fill=hl_fill7)
    )

    # ── SHEET 8: All 256 Scenarios (By Person) ────────────────────────────────
    ws8 = wb.create_sheet("All 256 Scenarios (By Person)")
    ws8.row_dimensions[1].height = 35
    hdr(ws8, 1, 1, "Scenario", 9); hdr(ws8, 1, 2, "KenPom Prob %", 13)
    for gi, g in enumerate(s16_info):
        t1 = team_name_map.get(g['team_1'],''); t2 = team_name_map.get(g['team_2'],'')
        hdr(ws8, 1, gi+3, f"Slot {g['slot_id']}\n{t1} vs {t2}", 18)
    pcs = len(s16_info)+3
    for pi2, pi in enumerate(person_sort_order):
        creator = creators[pi]
        hdr(ws8, 1, pcs+pi2, f"{creator}\n({round(person_baseline[pi]*100,1)}% base)", 20)
    for di, pi in enumerate(prob_order):
        perm = all_permutations[pi]; row_idx = di+2
        fill = ALT_FILL if row_idx % 2 == 0 else BLK_FILL
        cel_thick(ws8, row_idx, 1, pi+1, fill, bold=True)
        cel_thick(ws8, row_idx, 2, round(float(scenario_probs[pi])*100,4), fill, fmt='0.0000"%"')
        for gi, game in enumerate(s16_info):
            sid = game['slot_id']; forced = bool(forced_outcomes and sid in forced_outcomes)
            if forced:
                wname = forced_names.get(forced_outcomes[sid], team_name_map.get(forced_outcomes[sid],''))
                cell_fill = PatternFill("solid", start_color="FFD700")
                t_color = "000000" if get_team_color(wname) in ("FFCB05","FFCD00","CEB888") else get_team_color(wname)
            else:
                wid = game['team_1'] if perm[gi] == 0 else game['team_2']
                wname = team_name_map.get(wid,'')
                cell_fill = BLK_FILL
                t_color = get_team_color(wname)
            c = ws8.cell(row=row_idx, column=gi+3, value=wname)
            c.font = Font(name="Arial", size=9, bold=True, color=t_color)
            c.fill = cell_fill
            c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bdr_thick
        for pi2, pi_person in enumerate(person_sort_order):
            wp = round(all_person_results[pi][pi_person]*100, 2)
            base = person_baseline[pi_person]*100
            gf = gradient_fill(wp, base)
            fc = gradient_font_color(wp, base)
            cel_thick(ws8, row_idx, pcs+pi2, wp, gf, fmt='0.00"%"', color=fc)
    ws8.freeze_panes = "B2"
    ws8.auto_filter.ref = f"A1:{get_column_letter(pcs+len(creators)-1)}1"
    # Row highlight on click
    last_col8 = get_column_letter(pcs+len(creators)-1)
    hl_fill8  = PatternFill("solid", start_color="FFF2CC")
    ws8.conditional_formatting.add(
        f"A2:{last_col8}{257}",
        FormulaRule(formula=['ROW()=CELL("row")'], fill=hl_fill8)
    )

    wb.save(filename)
    print(f"Excel saved: {filename}")
    return filename

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main(n_sims, forced_outcomes=None, forced_names=None):
    forced_outcomes = forced_outcomes or {}
    forced_names    = forced_names    or {}
    print(f"\n{'='*65}")
    print(f"MARCH MADNESS POOL SIMULATOR v3.2")
    print(f"{'='*65}")
    print(f"Pure KenPom | Exhaustive 256 S16 | KenPom-weighted baseline | Person analysis")
    print(f"Sims/scenario: {n_sims:,} | Total: {256*n_sims:,}")
    if forced_outcomes:
        for sid, tid in forced_outcomes.items():
            print(f"  Forced: Slot {sid} → {forced_names.get(tid, str(tid))}")
    print()

    conn = get_connection()
    scores, picks, slots, strengths, teams = load_data(conn)
    conn.close()

    strength_map  = dict(zip(strengths['team_id'].astype(int), strengths['kenpom_net_rating']))
    team_name_map = dict(zip(teams['team_id'].astype(int), teams['team_name']))
    bracket_ids   = scores['bracket_id'].tolist()
    bracket_names = dict(zip(scores['bracket_id'], scores['bracket_name']))
    person_map    = build_person_map(scores)

    s16_info         = [{'slot_id': int(r['slot_id']), 'team_1': int(r['team_1_id']), 'team_2': int(r['team_2_id'])}
                        for _, r in slots.iterrows()]
    all_permutations = list(product([0, 1], repeat=8))
    all_results      = np.zeros((256, len(bracket_ids)))
    scenario_probs   = np.zeros(256)

    print(f"Running 256 scenarios...")
    for perm_idx, perm in enumerate(all_permutations):
        if (perm_idx+1) % 64 == 0:
            print(f"  Scenario {perm_idx+1}/256 complete...")
        s16_forced = {}
        for gi, game in enumerate(s16_info):
            sid = game['slot_id']
            if not (forced_outcomes and sid in forced_outcomes):
                s16_forced[sid] = game['team_1'] if perm[gi] == 0 else game['team_2']
        scenario_probs[perm_idx] = scenario_kenpom_prob(perm, s16_info, strength_map, forced_outcomes)
        aw = simulate_vectorized(n_sims, s16_forced, strength_map, slots, forced_outcomes)
        ts = score_all_brackets(n_sims, aw, picks, scores, bracket_ids)
        all_results[perm_idx] = compute_win_probs(ts)

    scenario_probs = scenario_probs / scenario_probs.sum()
    baseline_pct   = all_results.T @ scenario_probs
    best_pct       = all_results.max(axis=0)
    worst_pct      = all_results.min(axis=0)

    # Person-level results
    all_person_results, creators = compute_person_win_probs_from_bracket_results(
        all_results, bracket_ids, person_map)
    person_baseline = all_person_results.T @ scenario_probs
    person_best     = all_person_results.max(axis=0)
    person_worst    = all_person_results.min(axis=0)
    person_sort_order = np.argsort(-person_baseline)

    # Game impacts
    game_impacts = []
    for gi, game in enumerate(s16_info):
        forced = bool(forced_outcomes and game['slot_id'] in forced_outcomes)
        t1m = np.array([p[gi] == 0 for p in all_permutations]); t2m = ~t1m
        if forced:
            avg_t1 = baseline_pct.copy(); avg_t2 = baseline_pct.copy()
        else:
            t1p = scenario_probs.copy(); t1p[t2m] = 0; t1p /= t1p.sum()
            t2p = scenario_probs.copy(); t2p[t1m] = 0; t2p /= t2p.sum()
            avg_t1 = all_results.T @ t1p; avg_t2 = all_results.T @ t2p
        game_impacts.append({
            'slot_id': game['slot_id'], 'team_1': game['team_1'], 'team_2': game['team_2'],
            'team_1_name': team_name_map.get(game['team_1'], str(game['team_1'])),
            'team_2_name': team_name_map.get(game['team_2'], str(game['team_2'])),
            'avg_if_t1_wins': avg_t1, 'avg_if_t2_wins': avg_t2,
            'swing': avg_t1-avg_t2, 'abs_swing': np.abs(avg_t1-avg_t2),
        })

    # Print results
    sort_order = np.argsort(-baseline_pct)
    print(f"\n{'='*65}\nBRACKET RESULTS\n{'='*65}")
    print(f"\n{'Rank':<5} {'Bracket':<35} {'Pts':>5} {'Win%':>8} {'Best%':>8} {'Worst%':>8}")
    print("-" * 72)
    for rank, idx in enumerate(sort_order, 1):
        bid = bracket_ids[idx]; name = bracket_names[bid]
        pts = int(scores[scores['bracket_id'] == bid]['current_score'].iloc[0])
        print(f"{rank:<5} {name:<35} {pts:>5} {baseline_pct[idx]*100:>7.2f}% {best_pct[idx]*100:>7.2f}% {worst_pct[idx]*100:>7.2f}%")

    print(f"\n{'='*65}\nPERSON RESULTS\n{'='*65}")
    print(f"\n{'Rank':<5} {'Person':<22} {'Person Win%':>12} {'Best Bracket%':>14} {'Boost':>7}")
    print("-" * 65)
    bid_to_idx = {bid: i for i, bid in enumerate(bracket_ids)}
    for rank, pi in enumerate(person_sort_order, 1):
        creator  = creators[pi]
        bids     = person_map[creator]
        bid_idxs = [bid_to_idx[bid] for bid in bids if bid in bid_to_idx]
        best_b   = max(baseline_pct[i] for i in bid_idxs)*100 if bid_idxs else 0
        boost    = round(person_baseline[pi]*100 - best_b, 2)
        print(f"{rank:<5} {creator:<22} {person_baseline[pi]*100:>11.2f}% {best_b:>13.2f}% {boost:>+7.2f}%")

    build_excel(all_results, all_permutations, scenario_probs, s16_info,
                game_impacts, scores, bracket_ids, bracket_names,
                baseline_pct, best_pct, worst_pct, team_name_map,
                forced_outcomes, forced_names, person_map,
                all_person_results, person_baseline, person_best, person_worst, creators)
    print(f"\nDone!")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='March Madness Pool Simulator v3.2')
    parser.add_argument('--sims',  type=int, default=N_SIMS)
    parser.add_argument('--force', action='append', default=[], metavar='SLOT:TEAM')
    args = parser.parse_args()
    forced_outcomes = {}; forced_names = {}
    if args.force:
        conn = get_connection()
        for f in args.force:
            slot_str, team_name = f.split(':', 1)
            slot_id = int(slot_str.strip()); team_name = team_name.strip()
            row = pd.read_sql(f"SELECT team_id FROM {SCHEMA}.Teams WHERE team_name = '{team_name}'", conn)
            if len(row) == 0:
                print(f"WARNING: Team '{team_name}' not found!")
            else:
                team_id = int(row.iloc[0]['team_id'])
                forced_outcomes[slot_id] = team_id; forced_names[team_id] = team_name
        conn.close()
    main(args.sims, forced_outcomes or None, forced_names or None)
